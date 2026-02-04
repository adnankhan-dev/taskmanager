from fastapi import APIRouter, Request, Depends, Query
from sqlalchemy.orm import Session
from datetime import date

from app.dependencies import get_db, get_current_user
from app.ui import templates
from app.models.task import Task
from app.models.user import User

router = APIRouter(prefix="/ui/reports")


@router.get("")
def reports(
    request: Request,
    status: str | None = Query(None),
    priority: str | None = Query(None),
    assigned_to: int | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    q = db.query(Task).filter(Task.archived == False)

    if status:
        q = q.filter(Task.status == status)
    if priority:
        q = q.filter(Task.priority == priority)
    if assigned_to:
        q = q.filter(Task.assigned_to_id == assigned_to)
    if start_date:
        q = q.filter(Task.final_deadline >= start_date)
    if end_date:
        q = q.filter(Task.final_deadline <= end_date)

    tasks = q.order_by(Task.final_deadline).all()
    users = db.query(User).all()

    return templates.TemplateResponse(
        "reports/list.html",
        {
            "request": request,
            "tasks": tasks,
            "users": users,
            "user": user
        }
    )
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.dependencies import get_db, get_current_user
from app.services.report_service import get_filtered_tasks
from app.models.user import User

router = APIRouter(prefix="/ui/reports")


@router.get("/tasks/excel")
def export_tasks_excel(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),

    status: str | None = None,
    priority: str | None = None,
    type_id: int | None = None,
    assigned_to_id: int | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
):
    filters = locals()
    tasks = get_filtered_tasks(db, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks Report"

    # ================= HEADER =================
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{user.department} – Task Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # ================= COLUMN HEADERS =================
    headers = ["Title", "Type", "Status", "Priority", "Assigned To", "Deadline"]
    ws.append([])
    ws.append(headers)

    for col in ws.iter_cols(min_row=3, max_row=3):
        for cell in col:
            cell.font = Font(bold=True)

    # ================= DATA =================
    for task in tasks:
        ws.append([
            task.title,
            task.type.name if task.type else "",
            task.status,
            task.priority,
            task.assigned_to.username if task.assigned_to else "",
            task.final_deadline.strftime("%Y-%m-%d"),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=tasks_report.xlsx"
        }
    )
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO


@router.get("/tasks/pdf")
def export_tasks_pdf(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),

    status: str | None = None,
    priority: str | None = None,
    type_id: int | None = None,
    assigned_to_id: int | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
):
    filters = locals()
    tasks = get_filtered_tasks(db, filters)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    elements = []

    # ================= HEADER =================
    elements.append(
        Paragraph(
            f"<b>{user.department} – Task Report</b>",
            styles["Title"]
        )
    )

    elements.append(Paragraph("<br/>", styles["Normal"]))

    # ================= TABLE =================
    data = [
        ["Title", "Type", "Status", "Priority", "Assigned To", "Deadline"]
    ]

    for t in tasks:
        data.append([
            t.title,
            t.type.name if t.type else "",
            t.status,
            t.priority,
            t.assigned_to.username if t.assigned_to else "",
            t.final_deadline.strftime("%Y-%m-%d"),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=tasks_report.pdf"
        }
    )
